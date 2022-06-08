# Every cell in a spreadsheet gets it's coordinates inverted. For example,
# sheet[3][5] turns to sheet[5][3] in new spreadsheet.


import openpyxl, os


sheetFile = r"C:\Coding Shit\myPrograms\Practice Projects\Chapter 12\cellInverter\spreadsheet.xlsx"
wb = openpyxl.load_workbook(sheetFile)
os.chdir(os.path.dirname(sheetFile))
sheet = wb.active
columns = []
 

for column in range(1, sheet.max_column + 1):
   cells = []
   for row in range(1, sheet.max_row + 1):
      cells.append(sheet.cell(row=row, column=column).value)
   columns.append([i for i in cells])
   cells = []

   
wbInverted = openpyxl.Workbook()
sheet1 = wbInverted.active
for col in columns:
   for cell in col:
      sheet1.cell(row=columns.index(col) + 1, column=col.index(cell) + 1).value = cell
wbInverted.save('InvertedSpreadSheet.xlsx')
