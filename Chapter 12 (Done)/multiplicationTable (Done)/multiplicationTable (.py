import openpyxl, os, sys
if len(sys.argv) < 2:
   print('multiplicationTable.py - Input number to create multiplication table')
else:
   os.chdir(r"C:\Coding Shit\myPrograms\Practice Projects\Chapter 12\multiplicationTable")
   wb = openpyxl.Workbook()
   sheet = wb.active
   for num in range(2, int(sys.argv[1]) + 1):
      sheet.cell(row=1, column=num).value = num - 1
      sheet.cell(row=num, column=1).value = num - 1
   for column in range(2, int(sys.argv[1]) + 1):
      for empty_cell in range(2, int(sys.argv[1]) + 1):
         sheet.cell(row=empty_cell, column=column).value = sheet.cell(row=empty_cell, column=1).value * sheet.cell(row=1, column=column).value
   wb.save('table_%s.xlsx' % (sys.argv[1]))
