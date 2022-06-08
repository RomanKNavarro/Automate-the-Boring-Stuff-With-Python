import os, openpyxl
from openpyxl.utils import get_column_letter


wb = openpyxl.load_workbook(r"C:\Coding Shit\myPrograms\Practice Projects\Chapter 12\textFilesToSpreadsheet\spreadsheet.xlsx")
sheet = wb.get_sheet_by_name('Sheet1')
directory = r"C:\Coding Shit\myPrograms\Practice Projects\Chapter 12\textFilesToSpreadsheet\Ooga booga (test files)"
os.chdir(directory)

text2 = [] 
for filename in os.listdir(directory):
   if filename.endswith('.txt'):
      textFile = open(filename)
      text = textFile.readlines()
      text2.append(text)

for column in range(2, (len(text2))):
   for row in range(2, len(text2)):       
      sheet.cell(row=row, column=column).value = 'african'
                    

                    
wb.save(r'C:\Coding Shit\myPrograms\Practice Projects\Chapter 12\textFilesToSpreadsheet\newSpreadsheet.xlsx')
