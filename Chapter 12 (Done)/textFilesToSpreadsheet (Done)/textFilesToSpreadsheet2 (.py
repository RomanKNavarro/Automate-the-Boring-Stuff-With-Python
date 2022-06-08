# After reading several text files, each of their contents are added to a single column, with one line of text per row.

import os, openpyxl
from openpyxl.utils import get_column_letter 
textFilesDir = r"C:\Coding Shit\myPrograms\Practice Projects\Chapter 12\textFilesToSpreadsheet\Ooga Booga"
os.chdir(textFilesDir)
fileContent = []


for filename in os.listdir(textFilesDir):
   if filename.endswith('.txt'):
      textFile = open(filename)
      fileContent.append(textFile.readlines())
      textFile.close()

      
wb = openpyxl.Workbook()
sheet = wb.active
current_column = 1


for text in fileContent:
   current_row = 1
   for line in text:
      sheet[get_column_letter(current_column) + str(current_row)] = line
      current_row += 1
   current_column += 1
   print('added text to column' + str(current_column-1))

   
wb.save(r"C:\Coding Shit\myPrograms\Practice Projects\Chapter 12\textFilesToSpreadsheet\spreadsheet2.xlsx")
