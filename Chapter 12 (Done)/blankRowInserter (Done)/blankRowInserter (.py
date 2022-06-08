# The command line should accept 3 arguments, the first 2 being integers
# and the third a spreadsheet filename, starting at row N (first integer),
# all other M row (second integer) after it should be blank.


import openpyxl, sys, os
if len(sys.argv) < 2:
   print('blankRowInserter.py - First: integer starting row, Second: integer for number of rows, Third: filename')
else:
   directory = os.path.dirname(sys.argv[3])
   os.chdir(directory)
   old_wb = openpyxl.load_workbook(sys.argv[3])
   new_wb = openpyxl.Workbook()
   sheet1 = old_wb.active
   sheet2 = new_wb.active
   first = []
   second = []


   for row in range(1, int(sys.argv[1]) + 1):
      cells = []
      for column in range(1, sheet1.max_column + 1):
         cells.append(sheet1.cell(row=row, column=column).value)
         first.append([i for i in cells])
         cells = []                                                              # add all rows before blank to 'first'
   for row in range(int(sys.argv[2]) - 1, sheet1.max_row + 1):
      cells = []
      for column in range(1, sheet1.max_column + 1):
         cells.append(sheet1.cell(row=row, column=column).value)
      second.append([i for i in cells])
      cells = []                                                                 # add all rows after blank to 'second'


   for row in first:
      for value in row:
         sheet2.cell(row=first.index(row) + 1, column=row.index(value) + 1).value = value        # add all values in 'first' to 'new_wb'
   for row in second:
      index = int(sys.argv[1]) + int(sys.argv[2]) + 1
      for value in row:
         sheet2.cell(row=second.index(row) + index, column=row.index(value) + 1).value = value   # add spacing, then all values in 'second'
          

   new_wb.save('Inserted.xlsx')
 















































