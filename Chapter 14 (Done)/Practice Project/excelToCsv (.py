import openpyxl, os, csv, shutil

os.chdir(r"C:/Coding Shit/myPrograms/ATBS Chapter 14/Practice Project/files")
os.makedirs("habenNicht", exist_ok=True)


for excelFile in os.listdir("."):
   if not excelFile.endswith('xlsx'):
      continue

   print('Converting file: ' + excelFile)
   wb = openpyxl.load_workbook(excelFile)
   
   for sheetname in wb.get_sheet_names():
      sheet = wb.get_sheet_by_name(sheetname)
      newFileName = str(os.path.splitext(excelFile)[0]) + '_' + sheetname + '.csv'
      outputFile = open(os.path.join('habenNicht', newFileName), 'w', newline='')
      csvWriter = csv.writer(outputFile)

   for rowNum in range(1, sheet.max_row + 1):
      rowData = []
         
      for colNum in range(1, sheet.max_column + 1):
         rowData.append(sheet.cell(row=rowNum, column=colNum).value)

      csvWriter.writerow(rowData)
      outputFile.close()
            
shutil.move("habenNicht", r"C:\Coding Shit\myPrograms\ATBS Chapter 14\Practice Project")
   
   
